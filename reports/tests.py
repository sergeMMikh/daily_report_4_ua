from datetime import timedelta
from unittest.mock import patch

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from .models import Report


class ReportViewsTests(TestCase):
    def test_reports_can_be_sorted_by_datetime(self):
        older = Report.objects.create(
            occurred_at=timezone.now() - timedelta(hours=1), description_ru="Старая", description_pt="Antiga"
        )
        newer = Report.objects.create(
            occurred_at=timezone.now(), description_ru="Новая", description_pt="Nova"
        )
        descending = list(self.client.get(reverse("report-list") + "?sort=desc").context["reports"])
        ascending = list(self.client.get(reverse("report-list") + "?sort=asc").context["reports"])
        self.assertEqual([report.pk for report in descending], [newer.pk, older.pk])
        self.assertEqual([report.pk for report in ascending], [older.pk, newer.pk])

    def test_language_switch_is_saved_in_session(self):
        response = self.client.post(reverse("set-language"), {"language": "pt", "next": "/"})
        self.assertRedirects(response, "/")
        self.assertEqual(self.client.session["interface_language"], "pt")
        response = self.client.get(reverse("report-list"))
        self.assertContains(response, "Registos recentes")

    def test_portuguese_interface_saves_without_russian_text(self):
        session = self.client.session
        session["interface_language"] = "pt"
        session.save()
        now = timezone.localtime()
        response = self.client.post(reverse("report-create"), {
            "date": now.date().isoformat(),
            "time": now.strftime("%H:%M"),
            "description_pt": "Trabalho realizado diretamente em português.",
        })
        self.assertRedirects(response, reverse("report-list"))
        report = Report.objects.get()
        self.assertEqual(report.description_ru, "")
        self.assertEqual(report.description_pt, "Trabalho realizado diretamente em português.")

    def test_create_report(self):
        now = timezone.localtime()
        response = self.client.post(reverse("report-create"), {
            "date": now.date().isoformat(),
            "time": now.strftime("%H:%M"),
            "description_ru": "Проверил оборудование",
            "description_pt": "Verifiquei o equipamento",
        })
        self.assertRedirects(response, reverse("report-list"))
        self.assertEqual(Report.objects.count(), 1)

    @patch("reports.views.translate_to_portuguese", return_value="Trabalho concluído")
    def test_translate_endpoint(self, mocked_translate):
        response = self.client.post(reverse("translate"), {"text": "Работа выполнена"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["translation"], "Trabalho concluído")
        mocked_translate.assert_called_once_with("Работа выполнена", "ru")

    @patch("reports.views.translate_to_portuguese", return_value="Trabalho concluído")
    def test_english_interface_translates_from_english(self, mocked_translate):
        session = self.client.session
        session["interface_language"] = "en"
        session.save()
        response = self.client.post(reverse("translate"), {"text": "Work completed"})
        self.assertEqual(response.status_code, 200)
        mocked_translate.assert_called_once_with("Work completed", "en")

    def test_export_week_excludes_old_reports(self):
        Report.objects.create(
            occurred_at=timezone.now(), description_ru="Свежая", description_pt="Recente"
        )
        Report.objects.create(
            occurred_at=timezone.now() - timedelta(days=8), description_ru="Старая", description_pt="Antiga"
        )
        response = self.client.get(reverse("export-reports", args=["week"]))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        values = list(workbook.active.values)
        self.assertEqual(len(values), 2)
        self.assertEqual(values[0], ("Data", "Hora", "Descrição (PT)"))
        self.assertEqual(values[1][2], "Recente")

    def test_export_selected_year(self):
        Report.objects.create(
            occurred_at=timezone.datetime(2024, 6, 1, tzinfo=timezone.get_current_timezone()),
            description_ru="2024", description_pt="PT 2024",
        )
        Report.objects.create(
            occurred_at=timezone.datetime(2025, 6, 1, tzinfo=timezone.get_current_timezone()),
            description_ru="2025", description_pt="PT 2025",
        )
        response = self.client.get(reverse("export-reports", args=["year"]) + "?year=2024")
        workbook = load_workbook(filename=__import__("io").BytesIO(response.content))
        values = list(workbook.active.values)
        self.assertEqual([row[2] for row in values[1:]], ["PT 2024"])
        self.assertIn("daily-reports-year-2024.xlsx", response["Content-Disposition"])
