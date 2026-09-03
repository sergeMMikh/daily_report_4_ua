from datetime import timedelta
from io import BytesIO

from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .forms import ReportForm
from .i18n import LANGUAGES, TRANSLATIONS, current_language
from .models import Report
from .services import TranslationUnavailable, translate_to_portuguese


def report_list(request):
    form = ReportForm(language=current_language(request))
    sort = request.GET.get("sort", "desc")
    if sort not in {"asc", "desc"}:
        sort = "desc"
    ordering = "occurred_at" if sort == "asc" else "-occurred_at"
    reports = Report.objects.order_by(ordering)[:20]
    return render(request, "reports/index.html", {
        "form": form, "reports": reports, "sort": sort, "years": available_years()
    })


@require_POST
def create_report(request):
    form = ReportForm(request.POST, language=current_language(request))
    if form.is_valid():
        form.save()
        messages.success(request, TRANSLATIONS[current_language(request)]["saved"])
        return redirect("report-list")
    reports = Report.objects.order_by("-occurred_at")[:20]
    return render(request, "reports/index.html", {
        "form": form, "reports": reports, "sort": "desc", "years": available_years()
    }, status=400)


@require_POST
def translate(request):
    text_ui = TRANSLATIONS[current_language(request)]
    text = request.POST.get("text", "").strip()
    if not text:
        return JsonResponse({"error": text_ui["enter_text"]}, status=400)
    try:
        return JsonResponse({"translation": translate_to_portuguese(text, current_language(request))})
    except TranslationUnavailable as exc:
        return JsonResponse({"error": str(exc)}, status=503)
    except Exception:
        return JsonResponse({"error": text_ui["translation_failed"]}, status=502)


@require_POST
def set_language(request):
    language = request.POST.get("language", "")
    if language in LANGUAGES:
        request.session["interface_language"] = language
    return redirect(request.POST.get("next") or "report-list")


def available_years():
    current_year = timezone.localdate().year
    stored_years = Report.objects.dates("occurred_at", "year", order="DESC")
    return sorted({current_year, *(date.year for date in stored_years)}, reverse=True)


def export_reports(request, period):
    now = timezone.localtime()
    if period == "week":
        start = now - timedelta(days=7)
        title = "Последние 7 дней"
    elif period == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        title = "Текущий месяц"
    elif period == "year":
        try:
            selected_year = int(request.GET.get("year", now.year))
        except (TypeError, ValueError):
            return HttpResponse("Invalid year", status=400)
        if not 1900 <= selected_year <= 2100:
            return HttpResponse("Invalid year", status=400)
        start = now.replace(year=selected_year, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end = start.replace(year=selected_year + 1)
        title = str(selected_year)
    else:
        return HttpResponse("Неизвестный период", status=404)

    if period == "year":
        rows = Report.objects.filter(occurred_at__gte=start, occurred_at__lt=end).order_by("occurred_at")
    else:
        rows = Report.objects.filter(occurred_at__gte=start, occurred_at__lte=now).order_by("occurred_at")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily reports"
    sheet.append(["Data", "Hora", "Descrição (PT)"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="172033")
    for report in rows:
        local = timezone.localtime(report.occurred_at)
        sheet.append([local.date(), local.time().replace(tzinfo=None), report.description_pt])
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 11
    sheet.column_dimensions["C"].width = 85
    for row in sheet.iter_rows(min_row=2):
        row[0].number_format = "DD.MM.YYYY"
        row[1].number_format = "HH:MM"
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    suffix = selected_year if period == "year" else f"{now:%Y-%m-%d}"
    response["Content-Disposition"] = f'attachment; filename="daily-reports-{period}-{suffix}.xlsx"'
    response["X-Report-Period"] = title
    return response
