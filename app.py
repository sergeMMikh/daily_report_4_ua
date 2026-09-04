import io
import json
import os
import secrets
import socket
import sys
import threading
import webbrowser
from datetime import datetime, timedelta
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

# PyInstaller's windowed mode sets these streams to None. Some HTTP/OpenAI
# dependencies expect file-like streams during their first lazy import.
if getattr(sys, "frozen", False):
    if sys.stdout is None:
        sys.stdout = open(os.devnull, "w", encoding="utf-8")
    if sys.stderr is None:
        sys.stderr = open(os.devnull, "w", encoding="utf-8")

from flask import Flask, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from storage import JsonReportStore
from translator import api_key, translate_to_portuguese


LISBON = ZoneInfo("Europe/Lisbon")
APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
DATA_PATH = APP_DIR / "reports.json"
CONFIG_PATH = APP_DIR / "config.json"
PORT = 8765
DEFAULT_CONFIG = {"openai_api_key": "", "language": "pt"}
CONFIG_LOCK = threading.RLock()

TEXT = {
    "ru": {"subtitle":"Ежедневный отчёт о выполненной работе","date":"Дата","time":"Время","today":"Сегодня","now":"Сейчас","source":"Описание на русском","pt":"Версия на португальском","source_hint":"Опишите выполненную работу на русском...","pt_hint":"Перевод появится здесь...","translate":"Перевести →","save":"✓ Сохранить","recent":"Последние записи","new":"Сначала новые","old":"Сначала старые","edit":"Редактировать","delete":"Удалить","export":"Экспорт Excel","week":"Последние 7 дней","month":"Текущий месяц","year":"Текущий год","choose_year":"Выберите год","export_year":"Выгрузить год","empty":"Записей пока нет.","saved":"Отчёт сохранён.","updated":"Запись обновлена.","deleted":"Запись удалена.","edit_title":"Редактирование записи","cancel":"Отмена","confirm_delete":"Удалить эту запись?","translating":"Перевожу…","ready":"Перевод готов.","translation_unavailable":"Автоматический перевод отключён: добавьте openai_api_key в config.json.","description":"Описание"},
    "pt": {"subtitle":"Relatório diário das atividades realizadas","date":"Data","time":"Hora","today":"Hoje","now":"Agora","source":"Descrição","pt":"Descrição","source_hint":"","pt_hint":"Descreva o trabalho realizado...","translate":"Traduzir →","save":"✓ Guardar","recent":"Registos recentes","new":"Mais recentes","old":"Mais antigos","edit":"Editar","delete":"Eliminar","export":"Exportar para Excel","week":"Últimos 7 dias","month":"Mês atual","year":"Ano atual","choose_year":"Selecionar ano","export_year":"Exportar ano","empty":"Ainda não existem registos.","saved":"Relatório guardado.","updated":"Registo atualizado.","deleted":"Registo eliminado.","edit_title":"Editar registo","cancel":"Cancelar","confirm_delete":"Eliminar este registo?","translating":"A traduzir…","ready":"Tradução concluída.","translation_unavailable":"A tradução automática está desativada: adicione openai_api_key ao config.json.","description":"Descrição"},
    "en": {"subtitle":"Daily report of completed activities","date":"Date","time":"Time","today":"Today","now":"Now","source":"English description","pt":"Portuguese version","source_hint":"Describe the completed work in English...","pt_hint":"The translation will appear here...","translate":"Translate →","save":"✓ Save","recent":"Recent entries","new":"Newest first","old":"Oldest first","edit":"Edit","delete":"Delete","export":"Export to Excel","week":"Last 7 days","month":"Current month","year":"Current year","choose_year":"Select year","export_year":"Export year","empty":"No entries yet.","saved":"Report saved.","updated":"Entry updated.","deleted":"Entry deleted.","edit_title":"Edit entry","cancel":"Cancel","confirm_delete":"Delete this entry?","translating":"Translating…","ready":"Translation ready.","translation_unavailable":"Automatic translation is disabled: add openai_api_key to config.json.","description":"Description"},
}

app = Flask(__name__)
app.secret_key = os.getenv("DAILY_REPORT_SECRET", secrets.token_hex(32))
store = JsonReportStore(DATA_PATH)


def load_config():
    with CONFIG_LOCK:
        if CONFIG_PATH.exists():
            try:
                config = json.loads(CONFIG_PATH.read_text(encoding="utf-8-sig"))
                if not isinstance(config, dict):
                    config = {}
            except (json.JSONDecodeError, OSError):
                config = {}
        else:
            config = {}

        normalized = {**DEFAULT_CONFIG, **config}
        if normalized.get("language") not in TEXT:
            normalized["language"] = DEFAULT_CONFIG["language"]
        if not isinstance(normalized.get("openai_api_key"), str):
            normalized["openai_api_key"] = ""

        if normalized != config or not CONFIG_PATH.exists():
            temporary = CONFIG_PATH.with_suffix(CONFIG_PATH.suffix + ".tmp")
            temporary.write_text(
                json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8"
            )
            os.replace(temporary, CONFIG_PATH)
        return normalized


def save_language(selected):
    with CONFIG_LOCK:
        config = load_config()
        config["language"] = selected
        temporary = CONFIG_PATH.with_suffix(CONFIG_PATH.suffix + ".tmp")
        temporary.write_text(
            json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        os.replace(temporary, CONFIG_PATH)


def language():
    configured = load_config()["language"]
    selected = session.get("language", configured)
    return selected if selected in TEXT else configured


def parse_occurred(date_value, time_value):
    return datetime.strptime(f"{date_value} {time_value}", "%Y-%m-%d %H:%M").replace(tzinfo=LISBON)


def sorted_reports(order):
    return sorted(store.all(), key=lambda item: item["occurred_at"], reverse=order != "asc")


def years():
    return sorted({datetime.now(LISBON).year, *(datetime.fromisoformat(r["occurred_at"]).year for r in store.all())}, reverse=True)


@app.context_processor
def template_context():
    lang = language()
    return {"t": TEXT[lang], "language": lang}


@app.get("/")
def index():
    order = request.args.get("sort", "desc")
    if order not in {"asc", "desc"}:
        order = "desc"
    now = datetime.now(LISBON)
    return render_template("index.html", reports=sorted_reports(order), sort=order, now=now, years=years())


@app.post("/language")
def set_language():
    selected = request.form.get("language")
    if selected in TEXT:
        session["language"] = selected
        save_language(selected)
    return redirect(request.form.get("next") or url_for("index"))


@app.post("/reports")
def create_report():
    lang = language()
    try:
        occurred = parse_occurred(request.form["date"], request.form["time"])
    except (KeyError, ValueError):
        abort(400)
    source = request.form.get("description_source", "").strip()
    portuguese = request.form.get("description_pt", "").strip()
    if lang == "pt":
        source = ""
    if not portuguese or (lang != "pt" and not source):
        abort(400)
    store.add({"id": str(uuid4()), "occurred_at": occurred.isoformat(), "source_language": lang, "description_source": source, "description_pt": portuguese, "created_at": datetime.now(LISBON).isoformat(), "updated_at": datetime.now(LISBON).isoformat()})
    flash(TEXT[lang]["saved"])
    return redirect(url_for("index"))


@app.route("/reports/<report_id>/edit", methods=["GET", "POST"])
def edit_report(report_id):
    report = store.get(report_id)
    if not report:
        abort(404)
    if request.method == "POST":
        try:
            occurred = parse_occurred(request.form["date"], request.form["time"])
        except (KeyError, ValueError):
            abort(400)
        source = request.form.get("description_source", "").strip()
        portuguese = request.form.get("description_pt", "").strip()
        if not portuguese:
            abort(400)
        store.update(report_id, {"occurred_at": occurred.isoformat(), "description_source": source, "description_pt": portuguese, "updated_at": datetime.now(LISBON).isoformat()})
        flash(TEXT[language()]["updated"])
        return redirect(url_for("index"))
    occurred = datetime.fromisoformat(report["occurred_at"])
    return render_template("edit.html", report=report, occurred=occurred)


@app.post("/reports/<report_id>/delete")
def delete_report(report_id):
    if not store.delete(report_id):
        abort(404)
    flash(TEXT[language()]["deleted"])
    return redirect(url_for("index"))


@app.post("/translate")
def translate():
    text = request.form.get("text", "").strip()
    if not text:
        return jsonify(error="Empty text"), 400
    if not api_key(CONFIG_PATH):
        return jsonify(available=False, error=TEXT[language()]["translation_unavailable"])
    try:
        return jsonify(available=True, translation=translate_to_portuguese(text, language(), CONFIG_PATH))
    except Exception as exc:
        return jsonify(error=str(exc)), 502


@app.get("/export/<period>")
def export(period):
    now = datetime.now(LISBON)
    if period == "week":
        start, end, suffix = now - timedelta(days=7), now, now.date().isoformat()
    elif period == "month":
        start, end, suffix = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0), now, now.date().isoformat()
    elif period == "year":
        try:
            selected = int(request.args.get("year", now.year))
        except ValueError:
            abort(400)
        if not 1900 <= selected <= 2100:
            abort(400)
        start = datetime(selected, 1, 1, tzinfo=LISBON)
        end = datetime(selected + 1, 1, 1, tzinfo=LISBON)
        suffix = str(selected)
    else:
        abort(404)
    rows = [r for r in sorted_reports("asc") if start <= datetime.fromisoformat(r["occurred_at"]) < end]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Daily reports"
    sheet.append(["Data", "Hora", "Descrição (PT)"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="172033")
    for report in rows:
        occurred = datetime.fromisoformat(report["occurred_at"])
        sheet.append([occurred.date(), occurred.time().replace(tzinfo=None), report["description_pt"]])
    sheet.column_dimensions["A"].width, sheet.column_dimensions["B"].width, sheet.column_dimensions["C"].width = 14, 11, 85
    for row in sheet.iter_rows(min_row=2):
        row[0].number_format, row[1].number_format = "DD.MM.YYYY", "HH:MM"
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.freeze_panes = "A2"
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"daily-reports-{period}-{suffix}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def port_open():
    with socket.socket() as client:
        client.settimeout(0.2)
        return client.connect_ex(("127.0.0.1", PORT)) == 0


def main():
    load_config()
    if port_open():
        webbrowser.open(f"http://127.0.0.1:{PORT}/")
        return
    threading.Timer(1.0, webbrowser.open, args=(f"http://127.0.0.1:{PORT}/",)).start()
    from waitress import serve
    serve(app, host="127.0.0.1", port=PORT, threads=6)


if __name__ == "__main__":
    main()
